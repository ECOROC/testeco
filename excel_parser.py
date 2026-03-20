"""
parsers/excel_parser.py — Parsing des fichiers Excel DPGF
S'adapte automatiquement à n'importe quelle structure de colonnes.
"""

import logging
import re
from pathlib import Path
from typing import Optional

import openpyxl

from parsers.structure_detector import analyser_structure, StructureDPGF, _est_vide

logger = logging.getLogger(__name__)


def parser_excel(chemin: str, structure: StructureDPGF = None) -> list[dict]:
    """
    Parse un fichier Excel DPGF.

    Si une StructureDPGF est fournie (issue du fichier de référence),
    elle est utilisée directement.
    Sinon, la structure est détectée automatiquement.
    """
    chemin = Path(chemin)
    logger.info(f"Parsing Excel : {chemin.name}")

    wb = openpyxl.load_workbook(chemin, read_only=True, data_only=True)
    feuilles = _identifier_feuilles_dpgf(wb)

    if not feuilles:
        logger.warning(f"Aucune feuille DPGF dans {chemin.name}")
        wb.close()
        return []

    tous_postes = []
    for nom_feuille in feuilles:
        # Utiliser la structure fournie ou en détecter une
        if structure and structure.est_valide():
            struct_feuille = structure
        else:
            from parsers.structure_detector import _analyser_feuille
            struct_feuille = _analyser_feuille(wb[nom_feuille], nom_feuille)

        postes = _parser_feuille_avec_structure(wb[nom_feuille], nom_feuille, struct_feuille)
        tous_postes.extend(postes)
        logger.info(f"  '{nom_feuille}' → {len(postes)} poste(s)")

    wb.close()
    logger.info(f"Total : {len(tous_postes)} postes — {chemin.name}")
    return tous_postes


def parser_excel_avec_reference(chemin: str, chemin_reference: str) -> list[dict]:
    """
    Parse un fichier DPGF en utilisant la structure du fichier de référence.
    C'est la méthode à utiliser quand l'utilisateur fournit sa trame vierge.
    """
    logger.info(f"Lecture de la structure depuis la référence : {Path(chemin_reference).name}")
    structure_ref = analyser_structure(chemin_reference)

    if not structure_ref.est_valide():
        logger.warning("Structure référence invalide, détection automatique")
        return parser_excel(chemin)

    logger.info(f"Structure référence : {structure_ref}")
    return parser_excel(chemin, structure=structure_ref)


def _identifier_feuilles_dpgf(wb: openpyxl.Workbook) -> list[str]:
    """Identifie les feuilles contenant un DPGF."""
    feuilles_valides = []
    mots_exclusion = ["synthèse", "synthese", "récap", "recap",
                      "couverture", "sommaire", "note", "index"]

    for nom in wb.sheetnames:
        ws = wb[nom]
        nom_lower = nom.lower()

        # Exclure les feuilles de synthèse avec peu de contenu
        if any(m in nom_lower for m in mots_exclusion):
            continue

        # Vérifier qu'il y a du contenu substantiel
        nb_lignes = 0
        nb_numeriques = 0
        for row in ws.iter_rows(values_only=True):
            if any(v is not None for v in row):
                nb_lignes += 1
                nb_numeriques += sum(1 for v in row
                                     if isinstance(v, (int, float)) and v > 0)
            if nb_lignes > 30:
                break

        if nb_lignes >= 3 and nb_numeriques >= 2:
            feuilles_valides.append(nom)

    return feuilles_valides if feuilles_valides else wb.sheetnames[:1]


def _parser_feuille_avec_structure(ws, nom_feuille: str,
                                    structure: StructureDPGF) -> list[dict]:
    """Parse une feuille en utilisant la structure détectée."""
    mapping = structure.mapping
    debut   = structure.ligne_debut_data

    # Extraire le numéro de lot depuis le nom de feuille
    lot = _extraire_lot_depuis_nom(nom_feuille)

    postes = []
    lot_courant = lot
    lignes_lues = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < debut:
            continue

        if _est_vide(row):
            continue

        # Détecter les titres de section/lot (ligne sans prix)
        if _est_titre_section(row, mapping):
            nouveau_lot = _extraire_lot_depuis_ligne(row, mapping)
            if nouveau_lot:
                lot_courant = nouveau_lot
            continue

        # Ignorer les sous-totaux
        if _est_sous_total(row, mapping):
            continue

        poste = _extraire_poste(row, mapping, lot_courant, nom_feuille)
        if poste:
            postes.append(poste)
            lignes_lues += 1

        # Sécurité : arrêter après 2000 postes
        if lignes_lues > 2000:
            break

    return postes


def _extraire_poste(row: tuple, mapping: dict, lot: str,
                    nom_feuille: str) -> Optional[dict]:
    """Extrait un poste depuis une ligne selon le mapping."""
    def get(role):
        idx = mapping.get(role)
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        return v if v is not None else None

    def get_str(role):
        v = get(role)
        return str(v).strip() if v is not None else ""

    def get_float(role):
        v = get(role)
        if v is None: return None
        if isinstance(v, (int, float)): return float(v)
        s = re.sub(r"[^\d,.\-]", "", str(v)).replace(",", ".")
        try: return float(s) if s else None
        except: return None

    libelle = get_str("libelle")
    if not libelle or len(libelle) < 2:
        return None

    pu    = get_float("prix_unitaire")
    total = get_float("prix_total")

    # Quantité : essayer d'abord col entreprise, puis col référence en fallback
    qte = get_float("quantite")
    if qte is None:
        qte = get_float("quantite_ref")

    # Si toujours rien mais qu'on a un total et un PU, calculer la quantité
    if qte is None and pu and pu > 0 and total and total > 0:
        qte = round(total / pu, 3)

    if pu is None and total is None and qte is None:
        return None

    # Calculer le total si absent
    if total is None and pu is not None and qte is not None:
        total = round(pu * qte, 2)

    lot_final = lot or _extraire_lot_depuis_nom(nom_feuille) or "00"
    lot_nom   = _extraire_nom_lot(nom_feuille)

    return {
        "lot":           lot_final,
        "lot_nom":       lot_nom,
        "numero":        get_str("numero"),
        "libelle":       libelle,
        "unite":         get_str("unite"),
        "quantite":      qte,
        "prix_unitaire": pu,
        "prix_total":    total,
    }


def _est_titre_section(row: tuple, mapping: dict) -> bool:
    """Détecte si une ligne est un titre de section (pas un poste)."""
    # Une ligne de titre a généralement peu de valeurs non vides
    # et pas de valeurs numériques dans les colonnes prix
    col_pu = mapping.get("prix_unitaire")
    col_tot = mapping.get("prix_total")
    col_qte = mapping.get("quantite")

    for col in [col_pu, col_tot, col_qte]:
        if col is not None and col < len(row):
            v = row[col]
            if isinstance(v, (int, float)) and v > 0:
                return False

    # Vérifie qu'il y a un texte dans la colonne libellé
    col_lib = mapping.get("libelle", 1)
    if col_lib < len(row) and row[col_lib]:
        return True

    return False


def _est_sous_total(row: tuple, mapping: dict) -> bool:
    """Détecte les lignes de sous-total."""
    col_lib = mapping.get("libelle", 1)
    if col_lib < len(row) and row[col_lib]:
        texte = str(row[col_lib]).lower().strip()
        mots_st = ["total lot", "sous-total", "sous total",
                   "total ht", "total ttc", "report", "à reporter"]
        if any(m in texte for m in mots_st):
            return True
    return False


def _extraire_lot_depuis_nom(nom_feuille: str) -> str:
    """Extrait le numéro de lot depuis le nom d'une feuille Excel."""
    if not nom_feuille:
        return "00"
    match = re.search(r"lot\s*n?°?\s*(\d+|[A-Z]{1,3})", nom_feuille, re.IGNORECASE)
    if match:
        return match.group(1).zfill(2)
    match2 = re.search(r"(\d{1,2})\s*[-–]\s*\w", nom_feuille)
    if match2:
        return match2.group(1).zfill(2)
    return "00"


def _extraire_nom_lot(nom_feuille: str) -> str:
    """Extrait le nom descriptif du lot depuis le nom de feuille."""
    if not nom_feuille:
        return ""
    # "Lot N°02 GROS OEUVRE" → "GROS OEUVRE"
    clean = re.sub(r"lot\s*n?°?\s*\d*\s*[-–]?\s*", "", nom_feuille, flags=re.IGNORECASE)
    return clean.strip()


def _extraire_lot_depuis_ligne(row: tuple, mapping: dict) -> Optional[str]:
    """Cherche un numéro de lot dans une ligne de données."""
    for cell in row[:5]:
        if cell and isinstance(cell, str):
            match = re.search(r"lot\s*n?°?\s*(\d+|[A-Z]{1,3})\b",
                              cell.strip(), re.IGNORECASE)
            if match:
                return match.group(1).zfill(2)
    return None
