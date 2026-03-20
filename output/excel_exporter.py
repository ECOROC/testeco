"""
output/excel_exporter.py — Génération du fichier Excel comparatif multi-lots
1 feuille SYNTHÈSE GLOBALE + 1 feuille par lot
"""

import logging
from datetime import date
from pathlib import Path
from typing import Optional
import re

import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

from config import SEUIL_ALERTE_ECART

logger = logging.getLogger(__name__)

# ─── Palette ──────────────────────────────────────────────────────────────────
C_BLEU_FONCE  = "1F4E79"
C_BLEU_MOYEN  = "2E75B6"
C_BLEU_CLAIR  = "BDD7EE"
C_ROUGE       = "FFCCCC"
C_ROUGE_T     = "CC0000"
C_ORANGE      = "FFE5CC"
C_VERT_CLAIR  = "E8F5E9"
C_GRIS_CLAIR  = "F2F2F2"
C_GRIS_MOYEN  = "D6DCE4"
C_GRIS_FONCE  = "595959"
C_JAUNE_PALE  = "FFF3CD"
C_BLANC       = "FFFFFF"

def _fill(h): return PatternFill("solid", fgColor=h)
def _font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)

BORDER = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"),  bottom=Side(style="thin"))
AC = Alignment(horizontal="center", vertical="center")
AR = Alignment(horizontal="right",  vertical="center")
AL = Alignment(horizontal="left",   vertical="center", wrap_text=True)


# ═══════════════════════════════════════════════════════════════════════════════
# ENTRÉE PRINCIPALE
# ═══════════════════════════════════════════════════════════════════════════════

def generer_excel_comparatif(postes_reference, matchings_par_entreprise,
                             chemin_sortie, nom_operation="Comparatif DPGF"):
    chemin      = Path(chemin_sortie)
    entreprises = list(matchings_par_entreprise.keys())
    lots        = _extraire_lots(postes_reference)

    logger.info(f"Génération Excel : {len(lots)} lot(s) — {len(entreprises)} entreprise(s)")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_syn = wb.create_sheet("SYNTHÈSE GLOBALE")
    _synthese(ws_syn, postes_reference, matchings_par_entreprise,
              entreprises, lots, nom_operation)

    for lot in lots:
        postes_lot = [p for p in postes_reference if p.get("lot") == lot]
        if not postes_lot:
            continue
        nom_f = _nom_feuille(lot, postes_lot)
        ws = wb.create_sheet(nom_f)
        _lot(ws, postes_lot, matchings_par_entreprise, entreprises,
             lot, nom_f, nom_operation)
        logger.info(f"  '{nom_f}' : {len(postes_lot)} postes")

    wb.save(chemin)
    logger.info(f"Fichier Excel généré : {chemin}")
    return str(chemin)


# ═══════════════════════════════════════════════════════════════════════════════
# FEUILLE SYNTHÈSE GLOBALE
# ═══════════════════════════════════════════════════════════════════════════════

def _synthese(ws, postes_ref, matchings, entreprises, lots, nom_op):
    nb_ent  = len(entreprises)
    nb_cols = 3 + nb_ent + 4

    # Titre
    ws.merge_cells(f"A1:{get_column_letter(nb_cols)}1")
    c = ws["A1"]
    c.value = f"COMPARATIF DPGF — {nom_op.upper()}"
    c.font = _font(bold=True, color=C_BLANC, size=13)
    c.fill = _fill(C_BLEU_FONCE); c.alignment = AC
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{get_column_letter(nb_cols)}2")
    c = ws["A2"]
    c.value = (f"Synthèse globale — {date.today().strftime('%d/%m/%Y')} — "
               f"{len(entreprises)} entreprise(s) — {len(lots)} lot(s)")
    c.font = _font(italic=True, color=C_GRIS_FONCE, size=9)
    c.fill = _fill(C_BLEU_CLAIR); c.alignment = AC
    ws.row_dimensions[2].height = 14

    # En-têtes
    hdrs = ["Lot", "Intitulé", "Postes"] + \
           [_court(e) for e in entreprises] + \
           ["Offre min (€)", "Offre max (€)", "Écart", "Absents"]
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.border = BORDER; c.alignment = AC
        c.font = _font(bold=True, color=C_BLANC, size=9)
        c.fill = _fill(C_BLEU_FONCE if col <= 3
                       else C_BLEU_MOYEN if col <= 3+nb_ent
                       else C_GRIS_FONCE)
    ws.row_dimensions[3].height = 26

    totaux_g = {e: 0.0 for e in entreprises}
    absents_g = {e: 0 for e in entreprises}

    for i, lot in enumerate(lots, 4):
        ps_lot = [p for p in postes_ref if p.get("lot") == lot]
        intit  = _intitule(lot)

        _c(ws, i, 1, f"LOT {lot}", bold=True, align="center")
        _c(ws, i, 2, intit, bold=True)
        _c(ws, i, 3, len(ps_lot), align="center")

        totaux_lot = []; abs_lot = 0

        for j, ent in enumerate(entreprises):
            ml = [m for m in matchings[ent] if m["poste_reference"].get("lot") == lot]
            tot = sum(m.get("prix_total_retenu") or 0 for m in ml
                      if m.get("prix_total_retenu") and m.get("score_confiance") != "ABSENT")
            abs_ = sum(1 for m in ml if m.get("score_confiance") == "ABSENT")
            c = ws.cell(row=i, column=4+j, value=tot if tot > 0 else None)
            c.number_format = "#,##0.00 €"; c.border = BORDER; c.alignment = AR
            if tot > 0: totaux_lot.append(tot)
            totaux_g[ent] += tot; absents_g[ent] += abs_; abs_lot += abs_

        cs = 4 + nb_ent
        if totaux_lot:
            mn, mx = min(totaux_lot), max(totaux_lot)
            ec = (mx - mn) / mn if mn > 0 else 0
            _c(ws, i, cs,   mn, fmt="#,##0.00 €", fill=C_VERT_CLAIR, align="right")
            _c(ws, i, cs+1, mx, fmt="#,##0.00 €",
               fill=C_ROUGE if ec > SEUIL_ALERTE_ECART else None, align="right")
            c = ws.cell(row=i, column=cs+2, value=round(ec, 3))
            c.number_format = "0.0%"; c.border = BORDER; c.alignment = AC
            if ec > SEUIL_ALERTE_ECART:
                c.fill = _fill(C_ROUGE); c.font = _font(bold=True, color=C_ROUGE_T)
        else:
            for k in range(3): ws.cell(row=i, column=cs+k).border = BORDER

        c = ws.cell(row=i, column=cs+3, value=abs_lot if abs_lot > 0 else "")
        c.border = BORDER; c.alignment = AC
        if abs_lot > 0:
            c.fill = _fill(C_GRIS_MOYEN); c.font = _font(bold=True, color=C_ROUGE_T)

        if i % 2 == 0:
            for col in range(1, nb_cols+1):
                cell = ws.cell(row=i, column=col)
                if not cell.fill or cell.fill.fgColor.rgb in ("00000000","FFFFFFFF","FF000000"):
                    cell.fill = _fill(C_GRIS_CLAIR)

    # Total général
    rt = 4 + len(lots)
    ws.merge_cells(f"A{rt}:C{rt}")
    c = ws.cell(row=rt, column=1, value="TOTAL GÉNÉRAL TOUS LOTS")
    c.font = _font(bold=True, color=C_BLANC); c.fill = _fill(C_BLEU_FONCE)
    c.border = BORDER; c.alignment = AC

    tf = []
    for j, ent in enumerate(entreprises):
        tot = totaux_g[ent]
        _c(ws, rt, 4+j, tot if tot > 0 else None,
           fmt="#,##0.00 €", bold=True, fill=C_BLEU_FONCE, color=C_BLANC, align="right")
        if tot > 0: tf.append(tot)

    cs = 4 + nb_ent
    if tf:
        mn, mx = min(tf), max(tf)
        ec = (mx - mn) / mn if mn > 0 else 0
        _c(ws, rt, cs,   mn, fmt="#,##0.00 €", bold=True, fill=C_BLEU_FONCE, color=C_BLANC, align="right")
        _c(ws, rt, cs+1, mx, fmt="#,##0.00 €", bold=True, fill=C_BLEU_FONCE, color=C_BLANC, align="right")
        c = ws.cell(row=rt, column=cs+2, value=round(ec, 3))
        c.number_format = "0.0%"; c.font = _font(bold=True, color=C_BLANC)
        c.fill = _fill(C_BLEU_FONCE); c.border = BORDER; c.alignment = AC

    # Absents par entreprise
    ra = rt + 1
    ws.merge_cells(f"A{ra}:C{ra}")
    ws.cell(row=ra, column=1, value="Postes absents par entreprise").font = \
        _font(italic=True, color=C_GRIS_FONCE, size=9)
    for j, ent in enumerate(entreprises):
        nb = absents_g[ent]
        c = ws.cell(row=ra, column=4+j, value=nb if nb > 0 else "✓")
        c.alignment = AC
        c.font = _font(color=C_ROUGE_T if nb > 0 else "2E7D32", bold=(nb > 0))

    # Largeurs
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 9
    for j in range(nb_ent): ws.column_dimensions[get_column_letter(4+j)].width = 18
    for k in range(4):      ws.column_dimensions[get_column_letter(4+nb_ent+k)].width = 14
    ws.freeze_panes = "D4"


# ═══════════════════════════════════════════════════════════════════════════════
# FEUILLE PAR LOT
# ═══════════════════════════════════════════════════════════════════════════════

def _lot(ws, postes_lot, matchings, entreprises, lot, nom_f, nom_op):
    nb_ent  = len(entreprises)
    nb_cols = 5 + nb_ent + 5

    # Titre
    ws.merge_cells(f"A1:{get_column_letter(nb_cols)}1")
    c = ws["A1"]
    c.value = f"{nom_op} — {nom_f}"
    c.font = _font(bold=True, color=C_BLANC, size=12)
    c.fill = _fill(C_BLEU_FONCE); c.alignment = AC
    ws.row_dimensions[1].height = 26

    ws.merge_cells(f"A2:{get_column_letter(nb_cols)}2")
    c = ws["A2"]
    c.value = (f"{date.today().strftime('%d/%m/%Y')}  |  {len(postes_lot)} postes  |  "
               f"{nb_ent} entreprises  |  "
               "🔴 Écart >20%   🟠 Confiance moyenne   ⬜ Absent   🟡 Regroupé")
    c.font = _font(italic=True, color=C_GRIS_FONCE, size=8)
    c.fill = _fill(C_BLEU_CLAIR); c.alignment = AC
    ws.row_dimensions[2].height = 14

    # En-têtes
    hdrs = ["N°", "Désignation des ouvrages", "Unité", "Quantité", "Réf. PU (€)"] + \
           [f"{_court(e)}\nPU (€)" for e in entreprises] + \
           ["Min (€)", "Max (€)", "Moy. (€)", "Écart-type", "ALERTE"]
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.font = _font(bold=True, color=C_BLANC, size=9)
        c.fill = _fill(C_BLEU_FONCE if col <= 5
                       else C_BLEU_MOYEN if col <= 5+nb_ent
                       else C_GRIS_FONCE)
    ws.row_dimensions[3].height = 30
    ws.freeze_panes = ws.cell(row=4, column=3)

    total_ref = 0.0
    totaux_ent = {e: 0.0 for e in entreprises}

    for ri, poste in enumerate(postes_lot, 4):
        prix_ent = []

        _c(ws, ri, 1, poste.get("numero",""), align="center")
        _c(ws, ri, 2, poste.get("libelle",""), align="left")
        _c(ws, ri, 3, poste.get("unite",""),   align="center")
        _c(ws, ri, 4, poste.get("quantite"),   fmt="#,##0.000", align="right")
        _c(ws, ri, 5, poste.get("prix_unitaire"), fmt="#,##0.00", align="right")

        if poste.get("prix_total"): total_ref += poste["prix_total"]

        for j, ent in enumerate(entreprises):
            col = 6 + j
            m = _find_matching(matchings[ent], poste)
            cell = ws.cell(row=ri, column=col)
            cell.border = BORDER; cell.alignment = AR

            if not m or m["score_confiance"] == "ABSENT":
                cell.value = "ABSENT"
                cell.fill  = _fill(C_GRIS_MOYEN)
                cell.font  = _font(italic=True, color="888888", size=9)
                cell.alignment = AC

            else:
                pu = m.get("prix_unitaire_retenu")
                cell.value  = pu
                cell.number_format = "#,##0.00"
                sc = m["score_confiance"]
                if m["type"] == "regroupe":
                    cell.fill = _fill(C_JAUNE_PALE)
                    cell.comment = Comment("Poste regroupé", "DPGF Comparator")
                elif sc == "MOYEN":
                    cell.fill = _fill(C_ORANGE)
                elif sc == "FAIBLE":
                    cell.fill = _fill("FFF0E0")
                if pu is not None:
                    prix_ent.append((j, pu))

            if m and m.get("prix_total_retenu"):
                totaux_ent[ent] = totaux_ent.get(ent, 0) + m["prix_total_retenu"]

        # Stats
        cs = 6 + nb_ent
        vals = [pu for _, pu in prix_ent if pu is not None]

        if len(vals) >= 2:
            mn  = min(vals); mx = max(vals)
            moy = sum(vals) / len(vals)
            std = float(np.std(vals))
            ec  = (mx - mn) / moy if moy > 0 else 0
            al  = ec > SEUIL_ALERTE_ECART

            _c(ws, ri, cs,   mn,  fmt="#,##0.00", fill=C_VERT_CLAIR, align="right")
            _c(ws, ri, cs+1, mx,  fmt="#,##0.00",
               fill=C_ROUGE if al else None, align="right")
            _c(ws, ri, cs+2, round(moy, 2), fmt="#,##0.00", align="right")
            _c(ws, ri, cs+3, round(std, 2), fmt="#,##0.00", align="right")

            c_al = ws.cell(row=ri, column=cs+4, value="⚠ OUI" if al else "NON")
            c_al.border = BORDER; c_al.alignment = AC
            if al:
                c_al.fill = _fill(C_ROUGE)
                c_al.font = _font(bold=True, color=C_ROUGE_T)
            else:
                c_al.font = _font(color="2E7D32")

            # Colorier les valeurs aberrantes
            if al:
                for j, pu in prix_ent:
                    if moy > 0 and abs(pu - moy) / moy > SEUIL_ALERTE_ECART:
                        cell = ws.cell(row=ri, column=6+j)
                        if cell.value not in (None, "ABSENT"):
                            cell.fill = _fill(C_ROUGE)
        else:
            for k in range(5): _c(ws, ri, cs+k, None)

        ws.row_dimensions[ri].height = 18

    # Total lot
    rt = 4 + len(postes_lot)
    ws.merge_cells(f"A{rt}:D{rt}")
    c = ws.cell(row=rt, column=1, value=f"TOTAL — {nom_f}")
    c.font = _font(bold=True, color=C_BLANC)
    c.fill = _fill(C_BLEU_FONCE); c.border = BORDER; c.alignment = AC

    _c(ws, rt, 5, total_ref if total_ref > 0 else None,
       fmt="#,##0.00 €", bold=True, fill=C_BLEU_FONCE, color=C_BLANC, align="right")

    for j, ent in enumerate(entreprises):
        tot = totaux_ent.get(ent, 0)
        _c(ws, rt, 6+j, tot if tot > 0 else None,
           fmt="#,##0.00 €", bold=True, fill=C_BLEU_FONCE, color=C_BLANC, align="right")

    ws.row_dimensions[rt].height = 22

    # Largeurs
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 9
    ws.column_dimensions["D"].width = 11
    ws.column_dimensions["E"].width = 13
    for j in range(nb_ent): ws.column_dimensions[get_column_letter(6+j)].width = 15
    for k in range(5):      ws.column_dimensions[get_column_letter(6+nb_ent+k)].width = 13


# ═══════════════════════════════════════════════════════════════════════════════
# UTILITAIRES INTERNES
# ═══════════════════════════════════════════════════════════════════════════════

def _c(ws, row, col, value, fmt=None, bold=False, fill=None,
       color="000000", align="left", italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.border = BORDER
    if fmt:   c.number_format = fmt
    if fill:  c.fill = _fill(fill)
    c.font = _font(bold=bold, color=color, italic=italic)
    c.alignment = AC if align == "center" else AR if align == "right" else AL
    return c


def _find_matching(matchings, poste_ref):
    lib = poste_ref.get("libelle", "")
    lot = poste_ref.get("lot", "")
    for m in matchings:
        pr = m.get("poste_reference", {})
        if pr.get("libelle") == lib and pr.get("lot") == lot:
            return m
    for m in matchings:
        if m.get("poste_reference", {}).get("libelle") == lib:
            return m
    return None


def _extraire_lots(postes):
    vus = []
    for p in postes:
        lot = p.get("lot", "00")
        if lot not in vus:
            vus.append(lot)
    return sorted(vus)


def _nom_feuille(lot, postes_lot):
    nom = f"LOT {lot}"
    # Chercher un nom de lot dans les postes (champ lot_nom si présent)
    for p in postes_lot:
        if p.get("lot_nom"):
            suffixe = p["lot_nom"][:18].upper()
            return f"{nom} - {suffixe}"[:31]
    return nom[:31]


def _intitule(lot):
    """Intitulé lisible du lot — enrichi si on a un lot_nom."""
    return f"Lot {lot}"


def _court(nom_entreprise):
    """Nom court pour les en-têtes de colonnes."""
    n = re.sub(r"^Entreprise_?", "", nom_entreprise, flags=re.IGNORECASE)
    n = re.sub(r"_(BTP|Construction|Bâtiment|Batiment|SARL|SAS|SA)$",
               "", n, flags=re.IGNORECASE)
    return n.replace("_", " ").strip() or nom_entreprise
