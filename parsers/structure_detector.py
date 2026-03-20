"""
parsers/structure_detector.py

Détecte automatiquement la structure d'un fichier DPGF Excel
en analysant son contenu — indépendamment des noms de colonnes.

Principe :
1. On lit le fichier de référence vierge fourni par l'utilisateur
2. On identifie les colonnes par leur RÔLE (numéro, libellé, unité, qté, PU, total)
   en utilisant une combinaison de :
   - Analyse sémantique des en-têtes (si présents)
   - Analyse du contenu des cellules (type, format, position)
   - Heuristiques métier BTP
3. Ce mapping est ensuite réutilisé pour parser tous les fichiers entreprises
"""

import re
import logging
from typing import Optional

import openpyxl

logger = logging.getLogger(__name__)

# ─── Vocabulaire de reconnaissance des rôles de colonnes ──────────────────────
# Chaque entrée = (rôle, liste de fragments à chercher dans l'en-tête)
VOCABULAIRE = [
    ("numero", [
        "n°", "num", "numéro", "numero", "repère", "repere",
        "article", "ref", "poste", "art.", "rep.", "code",
        "chapitre", "item"
    ]),
    ("libelle", [
        "libellé", "libelle", "désignation", "designation",
        "description", "nature", "descriptif", "ouvrage",
        "travaux", "prestation", "fourniture", "intitulé",
        "intitule", "objet"
    ]),
    ("unite", [
        "unité", "unite", "u.", "unité de mesure", "unit"
    ]),
    ("quantite", [
        "quantité entreprise", "quantite entreprise",
        "qté entreprise", "qte entreprise",
        "quantité à renseigner", "quantité offre",
        "quantité", "quantite", "qté", "qte", "qt",
        "nombre", "q.", "nbre"
    ]),
    ("quantite_ref", [
        "quantité lacorps", "quantité maître", "quantité mo",
        "quantité projet", "quantité boe", "quantité be",
        "qté projet", "quantité dce", "quantité marché"
    ]),
    ("prix_unitaire", [
        "prix unitaire", "prix en €", "prix en eur",
        "p.u.", "pu", "pu ht", "p.u ht", "prix u.",
        "px unitaire", "tarif", "prix €", "prix ht",
        "coût unitaire", "cout unitaire", "prix/u"
    ]),
    ("prix_total", [
        "total en €", "total en eur", "prix total",
        "montant", "total ht", "montant ht", "total €",
        "montant €", "montant total", "total", "p.t.",
        "pt ht", "pt", "sous-total", "sous total"
    ]),
]


class StructureDPGF:
    """
    Représente la structure détectée d'un fichier DPGF.
    Contient le mapping colonnes → rôles et les paramètres de parsing.
    """

    def __init__(self):
        self.mapping: dict = {}          # {"numero": 0, "libelle": 1, ...}
        self.ligne_entete: int = 0       # Index de la ligne d'en-tête
        self.ligne_debut_data: int = 1   # Index de la première ligne de données
        self.nom_feuille: str = ""
        self.nb_colonnes_utiles: int = 0
        self.confiance: float = 0.0      # 0.0 à 1.0
        self.methode: str = ""           # "entetes" | "contenu" | "position"

    def est_valide(self) -> bool:
        """Un mapping est valide s'il contient au moins libellé + prix_unitaire."""
        return "libelle" in self.mapping and (
            "prix_unitaire" in self.mapping or "prix_total" in self.mapping
        )

    def col(self, role: str) -> Optional[int]:
        """Retourne l'index de colonne pour un rôle donné."""
        return self.mapping.get(role)

    def __repr__(self):
        return (f"StructureDPGF(méthode={self.methode}, "
                f"confiance={self.confiance:.0%}, "
                f"mapping={self.mapping})")


def analyser_structure(chemin_excel: str, nom_feuille: str = None) -> StructureDPGF:
    """
    Point d'entrée principal.
    Analyse un fichier Excel et retourne sa structure DPGF.

    Si nom_feuille est None, analyse la première feuille pertinente.
    """
    wb = openpyxl.load_workbook(chemin_excel, read_only=True, data_only=True)

    if nom_feuille and nom_feuille in wb.sheetnames:
        feuilles = [nom_feuille]
    else:
        feuilles = wb.sheetnames

    meilleure = None
    for nom in feuilles:
        structure = _analyser_feuille(wb[nom], nom)
        if structure.est_valide():
            if meilleure is None or structure.confiance > meilleure.confiance:
                meilleure = structure

    wb.close()

    if meilleure is None:
        logger.warning(f"Aucune structure DPGF détectée dans {chemin_excel}")
        meilleure = StructureDPGF()

    logger.info(f"Structure détectée : {meilleure}")
    return meilleure


def _analyser_feuille(ws, nom_feuille: str) -> StructureDPGF:
    """Analyse une feuille et retourne sa structure."""
    structure = StructureDPGF()
    structure.nom_feuille = nom_feuille

    data = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        data.append(row)
        if i >= 50:
            break

    if not data:
        return structure

    # ── Stratégie 1 : détection par en-têtes ──────────────────────────────────
    for i, row in enumerate(data[:20]):
        mapping = _matcher_entetes_semantique(row)
        if not mapping:
            continue

        # Compléter les colonnes implicites (numero et libelle souvent sans en-tête)
        mapping = _completer_mapping_implicite(mapping, data, i)

        # Préférer "Quantité Entreprises" sur "Quantité LACORPS"
        mapping = _preferer_quantite_entreprise(mapping, row)

        if len(mapping) >= 3 and ("libelle" in mapping or "prix_unitaire" in mapping):
            structure.mapping = mapping
            structure.ligne_entete = i
            structure.ligne_debut_data = _trouver_debut_apres_header(data, i)
            structure.methode = "entetes"
            structure.confiance = _calculer_confiance(mapping)
            if structure.est_valide():
                return structure

    # ── Stratégie 2 : détection par contenu ───────────────────────────────────
    mapping_contenu, ligne_debut = _detecter_par_contenu(data)
    if mapping_contenu and len(mapping_contenu) >= 3:
        mapping_contenu = _completer_mapping_implicite(mapping_contenu, data, 0)
        structure.mapping = mapping_contenu
        structure.ligne_debut_data = ligne_debut
        structure.ligne_entete = max(0, ligne_debut - 1)
        structure.methode = "contenu"
        structure.confiance = _calculer_confiance(mapping_contenu) * 0.85
        if structure.est_valide():
            return structure

    # ── Stratégie 3 : position standard ───────────────────────────────────────
    mapping_pos = _detecter_position_standard(data)
    if mapping_pos:
        structure.mapping = mapping_pos["mapping"]
        structure.ligne_debut_data = mapping_pos["debut"]
        structure.methode = "position"
        structure.confiance = 0.60
        return structure

    return structure


def _completer_mapping_implicite(mapping: dict, data: list, idx_header: int) -> dict:
    """
    Complète un mapping partiel avec les colonnes implicites.
    Dans beaucoup de DPGF, le N° et le Libellé n'ont pas d'en-tête explicite
    mais sont toujours en colonnes 0 et 1.
    """
    mapping = dict(mapping)

    # Trouver la première ligne de données après l'en-tête
    debut = _trouver_debut_apres_header(data, idx_header)
    lignes_data = [r for r in data[debut:debut+10] if not _est_vide(r)]

    if not lignes_data:
        return mapping

    # Si numero absent : col 0 contient-elle des numéros de postes ?
    if "numero" not in mapping:
        vals_col0 = [r[0] for r in lignes_data if r and r[0] is not None]
        if vals_col0:
            pattern = re.compile(r'^\s*[\d][\d\.\-\s]*\s*$')
            if sum(1 for v in vals_col0 if pattern.match(str(v))) >= len(vals_col0) * 0.4:
                mapping["numero"] = 0

    # Si libelle absent : col 1 contient-elle des textes longs ?
    if "libelle" not in mapping:
        cols_occupees = set(mapping.values())
        for col_idx in range(min(5, len(lignes_data[0]) if lignes_data else 5)):
            if col_idx in cols_occupees:
                continue
            vals = [r[col_idx] for r in lignes_data
                    if col_idx < len(r) and r[col_idx] is not None]
            strings = [str(v) for v in vals if isinstance(v, str) and len(str(v).strip()) > 5]
            if len(strings) >= len(vals) * 0.5:
                mapping["libelle"] = col_idx
                break

    return mapping


def _preferer_quantite_entreprise(mapping: dict, row: tuple) -> dict:
    """
    Si deux colonnes de quantité sont détectées (maître d'ouvrage + entreprise),
    préférer celle de l'entreprise.
    """
    mapping = dict(mapping)
    if "quantite" not in mapping:
        return mapping

    col_qte_actuelle = mapping["quantite"]
    if col_qte_actuelle >= len(row):
        return mapping

    cell_actuelle = str(row[col_qte_actuelle]).lower() if row[col_qte_actuelle] else ""

    # Si la quantité actuelle est celle du maître d'ouvrage, chercher celle de l'entreprise
    mots_maitrise = ["lacorps", "lacorp", "projet", "boe", "maître", "maitre",
                     "dce", "marché", "marche", "indicatif", "be "]
    mots_entreprise = ["entreprise", "renseigner", "offre", "soumission"]

    if any(m in cell_actuelle for m in mots_maitrise):
        # Chercher une colonne "entreprise" adjacente
        for offset in [1, -1, 2, -2]:
            col_candidate = col_qte_actuelle + offset
            if 0 <= col_candidate < len(row) and row[col_candidate] is not None:
                cell_candidate = str(row[col_candidate]).lower()
                if any(m in cell_candidate for m in mots_entreprise):
                    mapping["quantite"] = col_candidate
                    break

    return mapping


def _trouver_debut_apres_header(data: list, idx_header: int) -> int:
    """Trouve la première ligne de données réelles après l'en-tête."""
    pattern_num = re.compile(r'^\s*[\d][\d\.\-\s]*\s*$')

    for i in range(idx_header + 1, min(idx_header + 15, len(data))):
        row = data[i]
        vals = [v for v in row if v is not None and str(v).strip()]
        if len(vals) >= 2:
            premier = str(vals[0]).strip()
            if pattern_num.match(premier) and len(premier) < 20:
                return i
            # Ou bien : ligne avec des valeurs numériques
            if any(isinstance(v, (int, float)) and v > 0 for v in row):
                return i

    return idx_header + 1


def _matcher_entetes_semantique(row: tuple) -> dict:
    """
    Analyse une ligne et retourne le mapping rôle → index de colonne
    en cherchant les mots-clés du vocabulaire métier dans chaque cellule.
    """
    mapping = {}
    if not row:
        return mapping

    for idx, cell in enumerate(row):
        if cell is None:
            continue
        cell_str = str(cell).lower().strip()
        if not cell_str:
            continue

        for role, fragments in VOCABULAIRE:
            if role in mapping:
                continue  # déjà assigné
            for fragment in fragments:
                if fragment in cell_str:
                    # Priorité spéciale : "quantité entreprise" > "quantité" simple
                    if role == "quantite" and "quantite_ref" not in mapping:
                        if any(f in cell_str for f in ["entreprise", "renseigner", "offre"]):
                            mapping["quantite"] = idx
                        elif "quantite" not in mapping:
                            mapping["quantite"] = idx
                    else:
                        mapping[role] = idx
                    break

    # Si on a quantite_ref mais pas quantite, utiliser quantite_ref comme quantite
    if "quantite" not in mapping and "quantite_ref" in mapping:
        mapping["quantite"] = mapping.pop("quantite_ref")
    elif "quantite_ref" in mapping:
        del mapping["quantite_ref"]

    return mapping


def _detecter_par_contenu(data: list) -> tuple[dict, int]:
    """
    Détecte les colonnes en analysant le type et format du contenu.
    Retourne (mapping, index_premiere_ligne_donnees).

    Logique :
    - Colonne avec des numéros style "1.2.3" ou "2-1-1" → numero
    - Colonne avec des textes longs → libelle
    - Colonne avec des strings courts (m², ml, etc.) → unite
    - Colonnes numériques : discriminer par ordre de grandeur et position
    """
    # Trouver la première ligne qui ressemble à des données de postes
    debut = _trouver_debut_donnees(data)
    if debut is None:
        return {}, 0

    # Analyser 15 lignes de données maximum
    lignes_data = [row for row in data[debut:debut + 20]
                   if not _est_vide(row) and not _est_titre(row)][:15]

    if len(lignes_data) < 3:
        return {}, debut

    # Pour chaque colonne, collecter les valeurs non vides
    nb_cols = max(len(r) for r in lignes_data)
    # Limiter aux 30 premières colonnes pertinentes
    nb_cols = min(nb_cols, 30)

    profil_cols = {}
    for col_idx in range(nb_cols):
        valeurs = []
        for row in lignes_data:
            if col_idx < len(row) and row[col_idx] is not None:
                v = row[col_idx]
                if str(v).strip():
                    valeurs.append(v)
        if valeurs:
            profil_cols[col_idx] = _profil_colonne(valeurs)

    # Assigner les rôles
    mapping = {}
    col_scores = {}

    for col_idx, profil in profil_cols.items():
        scores = {}
        if profil["est_numero_poste"]:       scores["numero"]        = 0.9
        if profil["est_texte_long"]:         scores["libelle"]       = 0.85
        if profil["est_unite"]:              scores["unite"]         = 0.9
        if profil["est_quantite"]:           scores["quantite"]      = 0.7
        if profil["est_prix_unitaire"]:      scores["prix_unitaire"] = 0.75
        if profil["est_prix_total"]:         scores["prix_total"]    = 0.75
        col_scores[col_idx] = scores

    # Assigner en évitant les conflits (un rôle = une colonne)
    roles_prioritaires = ["numero", "libelle", "unite", "quantite",
                          "prix_unitaire", "prix_total"]
    for role in roles_prioritaires:
        meilleure_col = None
        meilleur_score = 0.5  # seuil minimum
        for col_idx, scores in col_scores.items():
            if col_idx in mapping.values():
                continue
            s = scores.get(role, 0)
            if s > meilleur_score:
                meilleur_score = s
                meilleure_col = col_idx
        if meilleure_col is not None:
            mapping[role] = meilleure_col

    return mapping, debut


def _profil_colonne(valeurs: list) -> dict:
    """Analyse le profil d'une colonne pour déterminer son rôle probable."""
    numeriques = [v for v in valeurs if isinstance(v, (int, float))]
    strings    = [str(v).strip() for v in valeurs if isinstance(v, str) and str(v).strip()]
    tous       = [str(v).strip() for v in valeurs if str(v).strip()]

    ratio_num = len(numeriques) / len(valeurs) if valeurs else 0
    ratio_str = len(strings)    / len(valeurs) if valeurs else 0

    # Numéros de postes : style "1.2.3" ou "2-1-1-1"
    pattern_num_poste = re.compile(r'^\d[\d\.\-]*\d?(\s*)$')
    est_numero_poste = (
        ratio_str > 0.5 and
        sum(1 for s in strings if pattern_num_poste.match(s)) / max(len(strings), 1) > 0.5
    )

    # Libellé : textes longs
    longueurs = [len(s) for s in strings]
    longueur_moy = sum(longueurs) / len(longueurs) if longueurs else 0
    est_texte_long = ratio_str > 0.6 and longueur_moy > 15 and not est_numero_poste

    # Unité : strings très courts (1-8 chars) qui ressemblent à des unités
    UNITES_CONNUES = {
        "m²","m2","m³","m3","ml","m","ml","kg","t","u","ens","ens.","forfait",
        "fft","pce","pièce","piece","so","h","j","semaine","mois",
        "m²/an","m3/h","ml/j","l","litre","nb","nbre","lot","k€"
    }
    est_unite = (
        ratio_str > 0.5 and
        sum(1 for s in strings if s.lower() in UNITES_CONNUES or len(s) <= 8) /
        max(len(strings), 1) > 0.6
        and longueur_moy < 10
        and not est_numero_poste
    )

    # Valeurs numériques
    valeurs_num = [v for v in numeriques if v > 0]
    if valeurs_num:
        min_v = min(valeurs_num)
        max_v = max(valeurs_num)
        moy_v = sum(valeurs_num) / len(valeurs_num)
    else:
        min_v = max_v = moy_v = 0

    # Quantité : nombres généralement petits et cohérents
    est_quantite = (
        ratio_num > 0.4 and len(valeurs_num) >= 2 and
        moy_v < 10000 and max_v < 100000 and
        not (moy_v > 500 and min_v > 50)  # Pas des prix
    )

    # PU : nombres moyens (prix unitaires BTP : 5€ à 50 000€)
    est_prix_unitaire = (
        ratio_num > 0.4 and len(valeurs_num) >= 2 and
        5 <= moy_v <= 100000
    )

    # Total : nombres généralement plus grands que PU
    est_prix_total = (
        ratio_num > 0.4 and len(valeurs_num) >= 2 and
        moy_v > 500
    )

    return {
        "est_numero_poste":  est_numero_poste,
        "est_texte_long":    est_texte_long,
        "est_unite":         est_unite,
        "est_quantite":      est_quantite,
        "est_prix_unitaire": est_prix_unitaire,
        "est_prix_total":    est_prix_total,
        "ratio_num":         ratio_num,
        "longueur_moy":      longueur_moy,
        "moy_valeur":        moy_v,
    }


def _detecter_position_standard(data: list) -> Optional[dict]:
    """
    Dernier recours : supposer un format standard
    (col 0=N°, 1=libellé, 2=unité, 3=qté, 4/5=PU, 5/6=total).
    """
    debut = _trouver_debut_donnees(data)
    if debut is None:
        return None

    # Vérifier que les colonnes 0 et 1 contiennent bien des données
    lignes_test = [row for row in data[debut:debut + 5] if not _est_vide(row)]
    if not lignes_test:
        return None

    # Chercher les colonnes numériques dans les premières colonnes
    cols_num = {}
    for row in lignes_test:
        for j, v in enumerate(row[:15]):
            if isinstance(v, (int, float)) and v > 0:
                cols_num[j] = cols_num.get(j, 0) + 1

    cols_triees = sorted(cols_num.keys())
    mapping = {"numero": 0, "libelle": 1}

    if len(cols_triees) >= 3:
        mapping["unite"]         = 2
        mapping["quantite"]      = cols_triees[0]
        mapping["prix_unitaire"] = cols_triees[1] if len(cols_triees) > 1 else cols_triees[0]
        mapping["prix_total"]    = cols_triees[2] if len(cols_triees) > 2 else cols_triees[-1]

    return {"mapping": mapping, "debut": debut}


def _trouver_debut_donnees(data: list) -> Optional[int]:
    """Trouve l'index de la première ligne qui ressemble à des données de postes."""
    pattern_num = re.compile(r'^\d[\d\.\-\s]*$')

    for i, row in enumerate(data[:30]):
        vals_non_vides = [v for v in row if v is not None and str(v).strip()]
        if len(vals_non_vides) < 2:
            continue

        premier = str(vals_non_vides[0]).strip()
        # Premier champ ressemble à un numéro de poste
        if pattern_num.match(premier) and len(premier) < 20:
            # ET il y a au moins une valeur numérique dans la ligne
            if any(isinstance(v, (int, float)) and v > 0 for v in row):
                return i

    return None


def _calculer_confiance(mapping: dict) -> float:
    """Calcule un score de confiance pour un mapping."""
    roles_importants = ["numero", "libelle", "unite", "quantite",
                        "prix_unitaire", "prix_total"]
    presents = sum(1 for r in roles_importants if r in mapping)
    return present / len(roles_importants) if (present := presents) > 0 else 0.0


def _est_vide(row: tuple) -> bool:
    return all(v is None or str(v).strip() == "" for v in row)


def _est_titre(row: tuple) -> bool:
    """Détecte les lignes de titre (peu de valeurs, pas de prix)."""
    vals = [v for v in row if v is not None and str(v).strip()]
    if len(vals) <= 2:
        return True
    nums = [v for v in vals if isinstance(v, (int, float)) and v > 0]
    return len(nums) == 0
