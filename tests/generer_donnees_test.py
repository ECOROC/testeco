"""
tests/generer_donnees_test.py — Génère les données de test réalistes
"""

import os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─── Postes de référence GO réalistes ─────────────────────────────────────────
POSTES_REFERENCE = [
    {"num": "1.1", "libelle": "Installation et repliement de chantier", "unite": "Forfait", "qte": 1, "pu": 8500},
    {"num": "1.2", "libelle": "Clôture provisoire de chantier - grillage H=2m", "unite": "ml", "qte": 120, "pu": 18},
    {"num": "2.1", "libelle": "Décapage de terre végétale sur 30 cm d'épaisseur", "unite": "m²", "qte": 850, "pu": 4.50},
    {"num": "2.2", "libelle": "Fouilles en rigole pour semelles filantes - terrain ordinaire", "unite": "m³", "qte": 145, "pu": 22},
    {"num": "2.3", "libelle": "Fouilles en puits pour semelles isolées - terrain ordinaire", "unite": "m³", "qte": 68, "pu": 28},
    {"num": "2.4", "libelle": "Remblai de fouilles compacté par couches de 30 cm", "unite": "m³", "qte": 95, "pu": 15},
    {"num": "3.1", "libelle": "Béton de propreté dosé à 150 kg/m³ - ép. 5 cm", "unite": "m³", "qte": 18, "pu": 95},
    {"num": "3.2", "libelle": "Semelles filantes béton armé - dosé à 350 kg/m³", "unite": "m³", "qte": 85, "pu": 185},
    {"num": "3.3", "libelle": "Semelles isolées béton armé - dosé à 350 kg/m³", "unite": "m³", "qte": 42, "pu": 210},
    {"num": "3.4", "libelle": "Longrines béton armé - section 25x40 cm", "unite": "ml", "qte": 180, "pu": 75},
    {"num": "4.1", "libelle": "Voiles béton armé - ép. 20 cm - dosé à 350 kg/m³", "unite": "m³", "qte": 125, "pu": 320},
    {"num": "4.2", "libelle": "Poteaux béton armé - section 30x30 cm", "unite": "ml", "qte": 95, "pu": 145},
    {"num": "4.3", "libelle": "Poutres béton armé - section 30x50 cm", "unite": "ml", "qte": 210, "pu": 165},
    {"num": "5.1", "libelle": "Plancher dalle pleine béton armé - ép. 20 cm", "unite": "m²", "qte": 420, "pu": 85},
    {"num": "5.2", "libelle": "Plancher hourdis - entrevous PSE - dalle de compression 5 cm", "unite": "m²", "qte": 380, "pu": 65},
    {"num": "6.1", "libelle": "Dallage béton armé - ép. 15 cm - treillis soudé", "unite": "m²", "qte": 650, "pu": 38},
    {"num": "6.2", "libelle": "Joint de fractionnement dallage - profilé aluminium", "unite": "ml", "qte": 85, "pu": 12},
    {"num": "7.1", "libelle": "Maçonnerie de blocs béton creux - 20 cm", "unite": "m²", "qte": 290, "pu": 45},
    {"num": "7.2", "libelle": "Enduit extérieur monocouche - finition grattée", "unite": "m²", "qte": 420, "pu": 22},
    {"num": "8.1", "libelle": "Chape ciment flottante - ép. 6 cm - isolation phonique", "unite": "m²", "qte": 680, "pu": 28},
]


def creer_excel_reference(chemin: Path):
    """Crée le DPGF de référence vierge (avec quantités, sans prix entreprise)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LOT 01 - GROS OEUVRE"

    # Style
    fill_header = PatternFill("solid", fgColor="1F4E79")
    font_blanc = Font(color="FFFFFF", bold=True, size=10)
    font_titre = Font(bold=True, size=14, color="1F4E79")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Titre
    ws.merge_cells("A1:G1")
    ws["A1"] = "DPGF - LOT 01 GROS OEUVRE - DOCUMENT DE RÉFÉRENCE"
    ws["A1"].font = font_titre
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:G2")
    ws["A2"] = "Construction de bureaux — Référence économiste"
    ws["A2"].font = Font(size=10, italic=True, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # En-têtes
    headers = ["N° Poste", "Désignation des ouvrages", "Unité", "Quantité", "Prix unitaire HT (€)", "Prix total HT (€)", "Observations"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.fill = fill_header
        c.font = font_blanc
        c.border = border
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[4].height = 35

    # Données
    for i, poste in enumerate(POSTES_REFERENCE, 5):
        ws.cell(row=i, column=1, value=poste["num"]).border = border
        lib = ws.cell(row=i, column=2, value=poste["libelle"])
        lib.border = border
        lib.alignment = Alignment(wrap_text=True)
        ws.cell(row=i, column=3, value=poste["unite"]).border = border
        c_qt = ws.cell(row=i, column=4, value=poste["qte"])
        c_qt.number_format = "#,##0.000"
        c_qt.border = border
        ws.cell(row=i, column=5, value="").border = border  # Vierge
        ws.cell(row=i, column=6, value="").border = border  # Vierge
        ws.cell(row=i, column=7, value="").border = border
        ws.row_dimensions[i].height = 18

    # Largeurs
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 20

    ws.freeze_panes = "A5"
    wb.save(chemin)
    print(f"✓ Référence créée : {chemin}")


def creer_excel_dupont(chemin: Path):
    """DPGF Dupont : postes standards, libellés proches."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DPGF GO"

    _entete_standard(ws, "ENTREPRISE DUPONT BTP — OFFRE DE PRIX")
    _headers_standard(ws)

    import random
    random.seed(42)

    for i, poste in enumerate(POSTES_REFERENCE, 5):
        pu = poste["pu"] * random.uniform(0.92, 1.08)
        total = round(pu * poste["qte"], 2)
        _ligne_poste(ws, i, poste["num"], poste["libelle"], poste["unite"], poste["qte"], round(pu, 2), total)

    _pied_tableau(ws, len(POSTES_REFERENCE) + 5)
    _colonnes_standard(ws)
    wb.save(chemin)
    print(f"✓ Dupont créé : {chemin}")


def creer_excel_martin(chemin: Path):
    """Martin : libellés différents, 2 postes manquants, 1 regroupement."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Offre Martin"

    _entete_standard(ws, "MARTIN CONSTRUCTION SARL — RÉPONSE APPEL D'OFFRES")
    _headers_standard(ws)

    import random
    random.seed(77)

    # Libellés reformulés
    libelles_martin = {
        "1.1": "Mise en place et dépose installation chantier — forfait",
        "1.2": "Clôture de chantier type palissade H=2.00m",
        "2.1": "Décapage et évacuation terre végétale e=30cm",
        "2.2": "Terrassement en tranchée pour fondations continues",
        "2.3": "Terrassement localisé pour fondations ponctuelles",
        # "2.4" ABSENT — poste manquant
        "3.1": "Béton maigre sous fondations dosage 150kg/m³",
        "3.2": "Semelles continues béton armé B35",
        # "3.3" ABSENT — poste manquant
        "3.4": "Longrines de fondation BA 25x40",
        # 4.1 et 4.2 REGROUPÉS en un seul poste chez Martin
        "4.1_regroupe": "Ouvrages verticaux béton armé (voiles + poteaux)",
        "4.3": "Poutres et linteaux béton armé",
        "5.1": "Dalle pleine béton armé ép. 20cm",
        "5.2": "Plancher à entrevous béton — dalle de compression 50mm",
        "6.1": "Dallage béton armé e=15cm avec treillis",
        "6.2": "Profilé de joint de fractionnement aluminium",
        "7.1": "Mur en blocs béton creux ép. 20 — appareillage courant",
        "7.2": "Enduit de façade monocouche gratté",
        "8.1": "Chape flottante ciment e=6cm sur isolant phonique",
    }

    postes_martin = [p for p in POSTES_REFERENCE
                     if p["num"] not in ["2.4", "3.3"]]  # 2 postes manquants

    ligne = 5
    for poste in postes_martin:
        num = poste["num"]

        if num == "4.1":
            # Regroupement 4.1 + 4.2
            p42 = next(p for p in POSTES_REFERENCE if p["num"] == "4.2")
            pu_regroupe = 280  # Prix global pour les 2 types d'ouvrages verticaux
            qt_regroupe = poste["qte"] + p42["qte"]
            total = round(pu_regroupe * qt_regroupe, 2)
            _ligne_poste(ws, ligne, "4.1/4.2", libelles_martin["4.1_regroupe"],
                         "m³", qt_regroupe, pu_regroupe, total)
            ligne += 1
            continue
        elif num == "4.2":
            continue  # Déjà regroupé avec 4.1

        libelle = libelles_martin.get(num, poste["libelle"])
        pu = poste["pu"] * random.uniform(0.88, 1.12)
        total = round(pu * poste["qte"], 2)
        _ligne_poste(ws, ligne, num, libelle, poste["unite"], poste["qte"], round(pu, 2), total)
        ligne += 1

    _pied_tableau(ws, ligne)
    _colonnes_standard(ws)
    wb.save(chemin)
    print(f"✓ Martin créé : {chemin}")


def creer_excel_vinci(chemin: Path):
    """Vinci : postes dans ordre différent, 1 prix avec écart > 30%."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Réponse Vinci"

    _entete_standard(ws, "VINCI BÂTIMENT — DÉCOMPOSITION DU PRIX GLOBAL ET FORFAITAIRE")
    _headers_standard(ws)

    import random
    random.seed(99)

    # Ordre aléatoire des postes
    postes_vinci = list(POSTES_REFERENCE)
    random.shuffle(postes_vinci)

    for i, poste in enumerate(postes_vinci, 5):
        pu = poste["pu"] * random.uniform(0.95, 1.05)

        # Poste 5.1 avec un écart important (+35%)
        if poste["num"] == "5.1":
            pu = poste["pu"] * 1.35

        total = round(pu * poste["qte"], 2)
        _ligne_poste(ws, i, poste["num"], poste["libelle"], poste["unite"], poste["qte"], round(pu, 2), total)

    _pied_tableau(ws, len(POSTES_REFERENCE) + 5)
    _colonnes_standard(ws)
    wb.save(chemin)
    print(f"✓ Vinci créé : {chemin}")


# ─── Fonctions utilitaires ─────────────────────────────────────────────────────

def _entete_standard(ws, titre: str):
    fill = PatternFill("solid", fgColor="E8F0FE")
    ws.merge_cells("A1:F1")
    ws["A1"] = titre
    ws["A1"].font = Font(bold=True, size=12, color="1F4E79")
    ws["A1"].fill = fill
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 25

    ws.merge_cells("A2:F2")
    ws["A2"] = "Opération : Construction de bureaux — Marché GO - Lot 01"
    ws["A2"].font = Font(size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    ws.merge_cells("A3:F3")
    ws["A3"] = "TVA applicable : 20% — Prix fermes et définitifs"
    ws["A3"].font = Font(size=9, color="666666")
    ws["A3"].alignment = Alignment(horizontal="center")


def _headers_standard(ws):
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True, size=10)
    headers = ["N°", "Désignation", "Unité", "Quantité", "P.U. HT (€)", "Montant HT (€)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.fill = fill
        c.font = font
        c.border = border
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[4].height = 30
    ws.freeze_panes = "A5"


def _ligne_poste(ws, row, num, libelle, unite, qte, pu, total):
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    ws.cell(row=row, column=1, value=num).border = border
    c = ws.cell(row=row, column=2, value=libelle)
    c.border = border
    c.alignment = Alignment(wrap_text=True)
    ws.cell(row=row, column=3, value=unite).border = border
    c_qt = ws.cell(row=row, column=4, value=qte)
    c_qt.number_format = "#,##0.000"
    c_qt.border = border
    c_pu = ws.cell(row=row, column=5, value=pu)
    c_pu.number_format = "#,##0.00"
    c_pu.border = border
    c_tot = ws.cell(row=row, column=6, value=total)
    c_tot.number_format = "#,##0.00"
    c_tot.border = border
    ws.row_dimensions[row].height = 20


def _pied_tableau(ws, ligne_total):
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="double"), bottom=Side(style="double"))
    fill = PatternFill("solid", fgColor="D6DCE4")
    ws.cell(row=ligne_total, column=2, value="TOTAL LOT 01 — GROS OEUVRE HT").font = Font(bold=True)
    ws.cell(row=ligne_total, column=2).fill = fill
    ws.cell(row=ligne_total, column=2).border = border
    # Formule de total
    ws.cell(row=ligne_total, column=6,
            value=f"=SUM(F5:F{ligne_total-1})").number_format = "#,##0.00"
    ws.cell(row=ligne_total, column=6).font = Font(bold=True)
    ws.cell(row=ligne_total, column=6).fill = fill
    ws.cell(row=ligne_total, column=6).border = border


def _colonnes_standard(ws):
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 18


def generer_tout():
    base = Path("test_data/AO_Construction_Bureaux")

    # Créer les dossiers
    dossiers = [
        base / "Entreprise_Dupont_BTP/DPGF",
        base / "Entreprise_Martin_Construction/DPGF",
        base / "Entreprise_Vinci_Batiment/DPGF",
    ]
    for d in dossiers:
        d.mkdir(parents=True, exist_ok=True)

    # Générer les fichiers
    creer_excel_dupont(base / "Entreprise_Dupont_BTP/DPGF/48291047.xlsx")
    creer_excel_martin(base / "Entreprise_Martin_Construction/DPGF/offre_finale.xlsx")
    creer_excel_vinci(base / "Entreprise_Vinci_Batiment/DPGF/reponse.xlsx")
    creer_excel_reference(Path("test_data/DPGF_Reference_GO_vierge.xlsx"))

    print("\n✅ Données de test générées dans test_data/")
    print("   Structure :")
    print("   test_data/")
    print("   ├── AO_Construction_Bureaux/")
    print("   │   ├── Entreprise_Dupont_BTP/DPGF/48291047.xlsx")
    print("   │   ├── Entreprise_Martin_Construction/DPGF/offre_finale.xlsx")
    print("   │   └── Entreprise_Vinci_Batiment/DPGF/reponse.xlsx")
    print("   └── DPGF_Reference_GO_vierge.xlsx")


if __name__ == "__main__":
    generer_tout()
